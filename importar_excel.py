    #Importaciones
import openpyxl
from openpyxl.styles import Font
import openpyxl.worksheet.dimensions 
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

def importar_excel(diccionario,tipo):
    #Asignacion de Variables
    biblioteca = openpyxl.Workbook()
    numero = 2
    negritas = Font(bold = True)
    borde_inferior = Border(bottom=Side(border_style="thin", color="000000"))
    centrado = Alignment(horizontal="center", vertical="center")

    #Asignacio de Negritas y Nombres
    primera = biblioteca["Sheet"]
    primera.title = " dia "

    primera["A1"].value = "ID Libro"
    primera["A1"].font = negritas
    primera["A1"].border = borde_inferior

    primera["B1"].value = "Titulo"
    primera["B1"].font = negritas
    primera["B1"].border = borde_inferior

    primera["C1"].value = "Autor"
    primera["C1"].font = negritas
    primera["C1"].border = borde_inferior

    primera["D1"].value = "Genero"
    primera["D1"].font = negritas
    primera["D1"].border = borde_inferior

    primera["E1"].value = "ISBN"
    primera["E1"].font = negritas
    primera["E1"].border = borde_inferior

    primera["F1"].value = "Fecha Publicacion"
    primera["F1"].font = negritas
    primera["F1"].border = borde_inferior

    primera["G1"].value = "Fecha Adquisicion"
    primera["G1"].font = negritas
    primera["G1"].border = borde_inferior

    #Ancho de Columnas
    primera.column_dimensions["A"].width = len(primera["A1"].value)
    primera.column_dimensions["B"].width = 10
    primera.column_dimensions["C"].width = 10
    primera.column_dimensions["D"].width = 10
    primera.column_dimensions["E"].width = 15
    primera.column_dimensions["F"].width = len(primera["F1"].value)
    primera.column_dimensions["G"].width = len(primera["G1"].value)

    for columna in range(1,8):
        for renglon in range(1,len(diccionario.keys())+2):
            primera.cell(row=renglon, column=columna).alignment = centrado

    #Ingreso de Datos
    lista_dicc = list(diccionario.keys())
    for fila in lista_dicc:
        primera.cell(row=numero, column=1).value = fila
        primera.cell(row=numero, column=2).value = diccionario[fila][0]
        primera.cell(row=numero, column=3).value = diccionario[fila][1]
        primera.cell(row=numero, column=4).value = diccionario[fila][2]
        primera.cell(row=numero, column=5).value = diccionario[fila][3]
        primera.cell(row=numero, column=6).value = diccionario[fila][4]
        primera.cell(row=numero, column=7).value = diccionario[fila][5]
        numero += 1
    


    if tipo == 0:
        typo = "Catalogo Completo"
    elif tipo == 1:
        typo = "Por Autor"
    elif tipo == 2:
        typo = "Por Genero"
    elif tipo == 3:
        typo = "Por Año de Publicacion"

    #Guardar Archivo
    biblioteca.save(f"Reporte de Libros {typo} .xlsx")
    print("Se creo el archivo correctamente")